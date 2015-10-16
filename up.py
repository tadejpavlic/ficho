from __future__ import division
import csv
import sys
import re
import datetime

from openpyxl import load_workbook

import logging
from sheetsync import Sheet
from local_settings import *

# Turn on logging so you can see what sheetsync is doing.

logging.getLogger('sheetsync').setLevel(logging.DEBUG)
logging.basicConfig()



# Xlsx data extraction & injection

fee = float(0.1696)

wb = load_workbook(filename = 'a.xlsx')
sheet_ranges = wb['Detailed Report']

for row in sheet_ranges.iter_rows():
	if row[1].value == 'Campaign':
		continue
	dt = row[0].value.strftime('%m/%d/%Y')
	campaign = row[1].value
	cost = row[2].value
	clicks = row[4].value
	impressions = row[5].value

	agg = { dt: {"Clicks" : clicks,
						"Impressions" : impressions,
						"CTR" : float(clicks/impressions),
						"Cost" : cost,
						"Platform Fee" : cost*fee,
						"Total Spend" : cost*fee + cost,
						"Avg. eCPC" : (cost*fee + cost)/clicks}}

	# Find or create a spreadsheet, then inject data.

	target = Sheet(credentials=creds,
	               document_key="1kTCqPid_y4npie6DuQAzX-p18F_OQZ1YUeKdyuN-m0M",
	               worksheet_name=campaign,
	               template_key="1kTCqPid_y4npie6DuQAzX-p18F_OQZ1YUeKdyuN-m0M",
	               key_column_headers=["Date"])

	target.inject(agg)

print "Spreadsheet created here: %s" % target.document_href